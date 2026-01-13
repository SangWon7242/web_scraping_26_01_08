from bs4 import BeautifulSoup
import requests # 웹사이트에 요청을 보내고 응답을 받는 라이브러리
import pandas as pd # 데이터 처리 및 엑셀 저장을 위한 라이브러리
from datetime import datetime # 오늘 날짜를 가져오기 위한 라이브러리
import os # 파일 존재 여부 확인을 위한 라이브러리
from openpyxl.styles import PatternFill, Border, Side, Font # 엑셀 스타일링을 위한 라이브러리


def get_unique_filename(base_path):
  """
  동일한 파일 명이 있다면 뒤에 번호를 붙여서 고유한 파일 이름 생성
  예: naver_news_20260113.xlsx -> naver_news_20260113_1.xlsx
  """
  # 파일이 존재하지 않으면 원래 이름 반환
  if not os.path.exists(base_path):
    return base_path
  
  # 파일 이름과 확장자 분리
  file_dir = os.path.dirname(base_path)
  file_name = os.path.basename(base_path)
  name, ext = os.path.splitext(file_name)
  
  # 번호를 붙여가며 고유한 파일 이름 찾기
  counter = 1
  while True:
    new_path = os.path.join(file_dir, f"{name}_{counter}{ext}")
    if not os.path.exists(new_path):
      return new_path
    counter += 1


def style_excel(file_path):
  """
  엑셀 파일에 스타일 적용: 제목 행 배경색, 모든 칸 테두리
  """
  from openpyxl import load_workbook
  
  # 엑셀 파일 열기
  wb = load_workbook(file_path)
  ws = wb.active
  
  # 제목 행 배경색 (파란색)
  header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
  # 제목 행 글자 색상 (흰색, 굵게)
  header_font = Font(color="FFFFFF", bold=True)
  # 테두리 스타일
  thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
  )
  
  # 모든 셀에 테두리 적용
  for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
      cell.border = thin_border
  
  # 제목 행(1행)에 배경색과 글자 스타일 적용
  for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
  
  # 열 너비 자동 조정 (가독성 향상)
  ws.column_dimensions['A'].width = 50  # 제목
  ws.column_dimensions['B'].width = 60  # 내용
  ws.column_dimensions['C'].width = 50  # 링크
  ws.column_dimensions['D'].width = 15  # 신문사
  
  # 저장
  wb.save(file_path)


# 1. requests : 원하는 웹사이트에 요청
url = "https://news.naver.com/section/105"

try:
  resp = requests.get(url, timeout=5)
  print(f"1. Requests: SUCCESS (Status Code: {resp.status_code})")
except Exception as e:
  print(f"1. Requests: FAILED ({e})")

if resp.status_code == 200:
  soup = BeautifulSoup(resp.text, 'lxml')
  
  # 페이지 헤드라인 (섹션 제목)
  section_title = soup.select_one(".sa_head_link")
  if section_title:
    print(f"섹션: {section_title.get_text(strip=True)}\n")
  
  # 모든 뉴스 항목 가져오기 (sa_text 클래스)
  news_items = soup.select(".sa_text")
  
  print(f"총 {len(news_items)}개의 뉴스를 찾았습니다.\n")
  print("=" * 80)
  
  # 각 뉴스 항목에서 정보 추출
  news_list = []
  
  for item in news_items:
    # 1. 링크 추출 (sa_text_title 클래스의 href 속성)
    link_element = item.select_one(".sa_text_title")
    link = link_element.get('href') if link_element else None
    
    # 2. 제목 추출 (sa_text_strong 클래스)
    title_element = item.select_one(".sa_text_strong")
    title = title_element.get_text(strip=True) if title_element else None
    
    # 3. 내용 추출 (sa_text_lede 클래스)
    content_element = item.select_one(".sa_text_lede")
    content = content_element.get_text(strip=True) if content_element else None
    
    # 4. 신문사 추출 (sa_text_info의 후손인 sa_text_press 클래스)
    news_corp_element = item.select_one(".sa_text_info .sa_text_press")
    news_corp = news_corp_element.get_text(strip=True) if news_corp_element else None
    
    # 딕셔너리로 정리 (조건5: 한글 키 사용)
    news_info = {
      "제목": title,
      "내용": content,
      "링크": link,
      "신문사": news_corp
    }
    
    news_list.append(news_info)
    
    # 각 뉴스 정보 출력
    print(f"\n📰 뉴스 {len(news_list)}")
    print(f"제목: {news_info['제목']}")
    print(f"신문사: {news_info['신문사']}")
    print(f"내용: {news_info['내용'][:50]}..." if news_info['내용'] else "내용: None")
    print(f"링크: {news_info['링크']}")
    print("-" * 80)
  
  print(f"\n✅ 총 {len(news_list)}개의 뉴스 정보 수집 완료!")
  
  # ======== 엑셀 파일로 저장 ========
  
  # excel_data 폴더가 없으면 생성
  os.makedirs("excel_data", exist_ok=True)
  
  # 오늘 날짜를 'YYYYMMDD' 형식으로 가져오기
  today = datetime.now().strftime("%Y%m%d")
  
  # -------- 조건2: 모든 뉴스 기사를 엑셀에 저장 --------
  # 저장할 파일 이름 생성 (조건4: 동일 파일명 있으면 번호 추가)
  base_file_name = f"excel_data/naver_news_{today}.xlsx"
  file_name = get_unique_filename(base_file_name)
  
  # news_list를 pandas DataFrame으로 변환
  df_all = pd.DataFrame(news_list)
  
  # 엑셀 파일로 저장 (index=False로 인덱스 번호 제외)
  df_all.to_excel(file_name, index=False)
  
  # 조건1: 엑셀 스타일 적용 (제목 행 배경색, 테두리)
  style_excel(file_name)
  
  # 저장 완료 메시지 출력
  print(f"\n📁 전체 뉴스 엑셀 파일 저장 완료: {file_name}")
  
  # -------- 조건3: 제목에 'AI'가 포함된 뉴스만 필터링해서 저장 --------
  # AI가 포함된 뉴스만 필터링 (대소문자 구분 없이)
  ai_news_list = [news for news in news_list if news['제목'] and 'AI' in news['제목'].upper()]
  
  if ai_news_list:
    # AI 뉴스 파일 이름 생성
    ai_base_file_name = f"excel_data/naver_news_AI_{today}.xlsx"
    ai_file_name = get_unique_filename(ai_base_file_name)
    
    # DataFrame으로 변환 후 저장
    df_ai = pd.DataFrame(ai_news_list)
    df_ai.to_excel(ai_file_name, index=False)
    
    # 스타일 적용
    style_excel(ai_file_name)
    
    print(f"📁 AI 뉴스 엑셀 파일 저장 완료: {ai_file_name} ({len(ai_news_list)}건)")
  else:
    print("\n⚠️ 제목에 'AI'가 포함된 뉴스가 없습니다.")